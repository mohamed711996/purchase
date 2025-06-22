"الشراء المقترح": st.column_config.NumberColumn(
                            "الشراء المقترح",
                            format="%.0f"
                        ),
                        "معدل دوران الكميات": st.column_config.NumberColumn(
                            "معدل دوران الكميات",
                            format="%.2f"
                        ),
                        "معدل دوران الفواتير": st.column_config.NumberColumn(
                            "معدل دوران الفواتير", 
                            format="%.2f"
                        ),
                        "عدد الفواتير": st.column_config.NumberColumn(
                            "عدد الفواتير",
                            format="%.0f"
                        ),
                        "متوسط الفواتير الشهرية": st.column_config.NumberColumn(import pandas as pd 
import streamlit as st 
from datetime import datetime, timedelta
import io
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# Configuration
st.set_page_config(
    page_title="نظام اقتراح المشتريات",
    page_icon="🛒",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better RTL support
st.markdown("""
<style>
    .main > div {
        direction: rtl;
        text-align: right;
    }
    .stSelectbox > div > div {
        direction: rtl;
    }
    .metric-card {
        background: linear-gradient(45deg, #667eea 0%, #764ba2 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin: 0.5rem 0;
    }
    .alert-high { background-color: #ff4444; color: white; padding: 0.5rem; border-radius: 5px; }
    .alert-medium { background-color: #ffaa00; color: white; padding: 0.5rem; border-radius: 5px; }
    .alert-low { background-color: #00aa00; color: white; padding: 0.5rem; border-radius: 5px; }
</style>
""", unsafe_allow_html=True)

# Load data from Excel files with improved error handling
@st.cache_data(ttl=300)  # Cache for 5 minutes
def load_data():
    """Load data from Excel files with enhanced error handling"""
    try:
        # Load with specific error handling for each file
        sales = pd.read_excel("sales_summary.xlsx")
        stock = pd.read_excel("Stocks.xlsx") 
        purchases = pd.read_excel("Purchase.xlsx")
        
        # Data validation - Updated column names
        required_sales_cols = ['Barcode', 'Year', 'Month', 'Quantity']
        required_stock_cols = ['Barcode', 'Name', 'Quantity On Hand']
        required_purchase_cols = ['Barcode', 'Date', 'purchase']
        
        # Check required columns
        missing_sales = [col for col in required_sales_cols if col not in sales.columns]
        missing_stock = [col for col in required_stock_cols if col not in stock.columns]
        missing_purchase = [col for col in required_purchase_cols if col not in purchases.columns]
        
        if missing_sales:
            st.error(f"أعمدة مفقودة في ملف المبيعات: {missing_sales}")
        if missing_stock:
            st.error(f"أعمدة مفقودة في ملف المخزون: {missing_stock}")
        if missing_purchase:
            st.error(f"أعمدة مفقودة في ملف المشتريات: {missing_purchase}")
            
        if missing_sales or missing_stock or missing_purchase:
            return None, None, None
            
        # Convert data types
        sales['Barcode'] = sales['Barcode'].astype(str)
        stock['Barcode'] = stock['Barcode'].astype(str)
        purchases['Barcode'] = purchases['Barcode'].astype(str)
        
        # Clean negative values
        sales['Quantity'] = sales['Quantity'].clip(lower=0)
        stock['Quantity On Hand'] = stock['Quantity On Hand'].clip(lower=0)
        purchases['purchase'] = purchases['purchase'].clip(lower=0)
        
        # Display column info for debugging
        st.write("📊 معلومات الأعمدة:")
        st.write(f"المبيعات: {list(sales.columns)}")
        st.write(f"المخزون: {list(stock.columns)}")
        st.write(f"المشتريات: {list(purchases.columns)}")
        
        return sales, stock, purchases
        
    except FileNotFoundError as e:
        st.error(f"❌ ملف غير موجود: {str(e)}")
        return None, None, None
    except Exception as e:
        st.error(f"❌ خطأ في تحميل البيانات: {str(e)}")
        return None, None, None

# Convert DataFrame to Excel bytes with formatting
def to_excel(df, filename_prefix="purchase_plan"):
    """Convert DataFrame to Excel with enhanced formatting"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='خطة الشراء')
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['خطة الشراء']
        
        # Apply formatting
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for col in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col)
            max_length = max(
                len(str(df.iloc[row, col-1])) for row in range(min(len(df), 100))
            )
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    return output.getvalue()

# Enhanced purchase plan generation
def generate_plan(sales, stock, purchases, target_month, target_year, safety_stock_days=30):
    """Generate purchase plan with improved algorithms"""
    
    # Calculate reference periods
    last_year = target_year - 1 
    prev_month = target_month - 1 if target_month > 1 else 12
    prev_year = target_year if target_month > 1 else target_year - 1
 
    # Previous month filter
    sales_prev_month = sales[
        (sales['Year'] == prev_year) & (sales['Month'] == prev_month)
    ]
    
    # Calculate 3 months from previous year
    months_last_year = []
    for i in range(1, 4):
        month = target_month - i
        if month > 0:
            months_last_year.append(month)
        else:
            months_last_year.append(12 + month)
    
    sales_last_year = sales[
        (sales['Year'] == last_year) & (sales['Month'].isin(months_last_year))
    ]
 
    # Combine sales data
    combined_sales = pd.concat([sales_prev_month, sales_last_year])
    
    if combined_sales.empty:
        st.warning("⚠️ لا توجد بيانات مبيعات للفترة المحددة")
        return pd.DataFrame()
    
    # Calculate months with actual sales per product and invoice count
    months_with_sales = combined_sales.groupby('Barcode').agg({
        'Month': 'nunique',
        'Quantity': 'sum',
        'Order Reference': 'nunique'  # عدد الفواتير الفريدة
    }).reset_index()
    months_with_sales.columns = ['Barcode', 'Months_With_Sales', 'Total_Sales_Period', 'Invoice_Count']
    
    # Calculate average monthly sales based on actual sales months
    months_with_sales['Average_Monthly_Sales'] = (
        months_with_sales['Total_Sales_Period'] / months_with_sales['Months_With_Sales']
    )
    
    # Calculate average invoices per month
    months_with_sales['Average_Monthly_Invoices'] = (
        months_with_sales['Invoice_Count'] / months_with_sales['Months_With_Sales']
    )
    
    # Process purchases data
    purchases = purchases.copy()
    purchases['Date'] = pd.to_datetime(purchases['Date'], errors='coerce')
    purchases = purchases.dropna(subset=['Date'])
    purchases['Year'] = purchases['Date'].dt.year
    purchases['Month'] = purchases['Date'].dt.month
    
    # Previous month purchases
    purchases_prev_month = purchases[
        (purchases['Year'] == prev_year) & (purchases['Month'] == prev_month)
    ]
    
    # Previous year purchases
    purchases_last_year = purchases[
        (purchases['Year'] == last_year) & (purchases['Month'].isin(months_last_year))
    ]
    
    # Combine purchase data
    combined_purchases = pd.concat([purchases_prev_month, purchases_last_year])
    
    # Summarize purchases
    if not combined_purchases.empty:
        purchases_summary = combined_purchases.groupby('Barcode').agg({
            'purchase': 'sum',
            'اسم المورد': lambda x: ', '.join(x.dropna().unique()) if 'اسم المورد' in combined_purchases.columns else 'غير محدد'
        }).reset_index()
        purchases_summary.columns = ['Barcode', 'Total_Purchases_Period', 'Suppliers']
    else:
        purchases_summary = pd.DataFrame(columns=['Barcode', 'Total_Purchases_Period', 'Suppliers'])
    
    # Merge with stock data
    df = stock.merge(months_with_sales, on='Barcode', how='left')
    df = df.merge(purchases_summary, on='Barcode', how='left')
    
    # Fill missing values
    numeric_columns = ['Total_Sales_Period', 'Average_Monthly_Sales', 'Months_With_Sales', 
                      'Invoice_Count', 'Average_Monthly_Invoices', 'Total_Purchases_Period']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = df[col].fillna(0)
    
    df['Suppliers'] = df['Suppliers'].fillna('غير محدد')
    
    # Calculate inventory metrics
    df['Average_Inventory'] = np.where(
        df['Total_Purchases_Period'] > 0,
        (df['Quantity On Hand'] + df['Total_Purchases_Period']) / 2,
        df['Quantity On Hand']
    )
    df['Average_Inventory'] = df['Average_Inventory'].replace(0, 1)
    
    # Calculate TWO turnover rates
    # 1. Traditional turnover rate (based on quantities)
    df['Quantity_Turnover_Rate'] = df['Total_Sales_Period'] / df['Average_Inventory']
    df['Quantity_Turnover_Rate'] = df['Quantity_Turnover_Rate'].round(2)
    
    # 2. Invoice-based turnover rate (considers frequency of sales)
    # This considers how often products are sold (invoice frequency)
    df['Invoice_Turnover_Rate'] = np.where(
        df['Average_Inventory'] > 0,
        df['Invoice_Count'] / (df['Average_Inventory'] / df['Average_Monthly_Sales'].replace(0, 1)),
        0
    )
    df['Invoice_Turnover_Rate'] = df['Invoice_Turnover_Rate'].round(2)
    
    # Enhanced turnover classification for both rates
    def classify_quantity_turnover(rate):
        if rate >= 6:
            return 'سريع جداً'
        elif rate >= 3:
            return 'سريع'
        elif rate >= 1.5:
            return 'متوسط'
        elif rate >= 0.5:
            return 'بطيء'
        else:
            return 'راكد'
    
    def classify_invoice_turnover(rate):
        if rate >= 8:
            return 'عالي التكرار'
        elif rate >= 4:
            return 'متوسط التكرار'
        elif rate >= 2:
            return 'منخفض التكرار'
        else:
            return 'نادر البيع'
    
    df['Quantity_Turnover_Classification'] = df['Quantity_Turnover_Rate'].apply(classify_quantity_turnover)
    df['Invoice_Turnover_Classification'] = df['Invoice_Turnover_Rate'].apply(classify_invoice_turnover)
    
    # Calculate safety stock
    df['Safety_Stock'] = (df['Average_Monthly_Sales'] * safety_stock_days) / 30
    
    # Enhanced purchase recommendation
    df['Days_Of_Stock'] = np.where(
        df['Average_Monthly_Sales'] > 0,
        (df['Quantity On Hand'] / df['Average_Monthly_Sales']) * 30,
        999
    )
    
    # Recommended purchase considers safety stock and lead time
    df['Recommended_Purchase'] = np.maximum(
        (df['Average_Monthly_Sales'] + df['Safety_Stock']) - df['Quantity On Hand'],
        0
    )
    
    # Priority classification
    def calculate_priority(row):
        if row['Days_Of_Stock'] <= 7:
            return 'عاجل جداً'
        elif row['Days_Of_Stock'] <= 15:
            return 'عاجل'
        elif row['Days_Of_Stock'] <= 30:
            return 'متوسط'
        else:
            return 'منخفض'
    
    df['Priority'] = df.apply(calculate_priority, axis=1)
    
    # Calculate cost metrics if available
    if 'Cost' in stock.columns:
        df['Total_Cost'] = df['Recommended_Purchase'] * df['Cost']
    else:
        df['Total_Cost'] = 0
    
    # Select and rename columns
    result_columns = [
        'Barcode', 'Name', 'Product Category/Complete Name', 'Quantity On Hand',
        'Total_Sales_Period', 'Invoice_Count', 'Months_With_Sales', 'Average_Monthly_Sales',
        'Average_Monthly_Invoices', 'Safety_Stock', 'Days_Of_Stock', 
        'Quantity_Turnover_Rate', 'Invoice_Turnover_Rate',
        'Quantity_Turnover_Classification', 'Invoice_Turnover_Classification', 
        'Priority', 'Recommended_Purchase', 'Total_Cost', 'Suppliers'
    ]
    
    # Filter columns that exist
    available_columns = [col for col in result_columns if col in df.columns]
    result_df = df[available_columns].copy()
    
    # Arabic column names
    arabic_names = {
        'Barcode': 'الباركود',
        'Name': 'اسم المنتج',
        'Product Category/Complete Name': 'فئة المنتج',
        'Quantity On Hand': 'الكمية المتاحة',
        'Total_Sales_Period': 'إجمالي المبيعات',
        'Invoice_Count': 'عدد الفواتير',
        'Months_With_Sales': 'عدد الشهور بمبيعات',
        'Average_Monthly_Sales': 'متوسط المبيعات الشهرية',
        'Average_Monthly_Invoices': 'متوسط الفواتير الشهرية',
        'Safety_Stock': 'مخزون الأمان',
        'Days_Of_Stock': 'أيام التغطية',
        'Quantity_Turnover_Rate': 'معدل دوران الكميات',
        'Invoice_Turnover_Rate': 'معدل دوران الفواتير',
        'Quantity_Turnover_Classification': 'تصنيف دوران الكميات',
        'Invoice_Turnover_Classification': 'تصنيف دوران الفواتير',
        'Priority': 'الأولوية',
        'Recommended_Purchase': 'الشراء المقترح',
        'Total_Cost': 'التكلفة الإجمالية',
        'Suppliers': 'الموردين'
    }
    
    result_df.rename(columns=arabic_names, inplace=True)
    
    # Round numeric columns
    numeric_cols = ['متوسط المبيعات الشهرية', 'متوسط الفواتير الشهرية', 'مخزون الأمان', 
                   'أيام التغطية', 'الشراء المقترح', 'معدل دوران الكميات', 'معدل دوران الفواتير']
    for col in numeric_cols:
        if col in result_df.columns:
            result_df[col] = result_df[col].round(2)
    
    return result_df

# Enhanced visualization functions
def create_turnover_charts(df):
    """Create both turnover analysis charts"""
    
    # Create subplots
    fig = make_subplots(
        rows=1, cols=2,
        subplot_titles=["معدل دوران الكميات", "معدل دوران الفواتير"],
        specs=[[{"type": "pie"}, {"type": "pie"}]]
    )
    
    # Quantity turnover chart
    quantity_counts = df['تصنيف دوران الكميات'].value_counts()
    fig.add_trace(
        go.Pie(
            labels=quantity_counts.index,
            values=quantity_counts.values,
            name="دوران الكميات",
            marker_colors=['#00CC96', '#19D3F3', '#FFA15A', '#FF6692', '#B6E880']
        ),
        row=1, col=1
    )
    
    # Invoice turnover chart
    invoice_counts = df['تصنيف دوران الفواتير'].value_counts()
    fig.add_trace(
        go.Pie(
            labels=invoice_counts.index,
            values=invoice_counts.values,
            name="دوران الفواتير",
            marker_colors=['#636EFA', '#EF553B', '#00CC96', '#AB63FA']
        ),
        row=1, col=2
    )
    
    fig.update_traces(textposition='inside', textinfo='percent+label')
    fig.update_layout(
        title_text="تحليل معدلات الدوران",
        font=dict(size=12),
        height=500,
        showlegend=True
    )
    
    return fig

def create_combined_analysis_chart(df):
    """Create combined analysis showing both turnover rates"""
    
    # Filter products with purchase recommendations
    df_filtered = df[df['الشراء المقترح'] > 0].copy()
    
    if df_filtered.empty:
        return None
    
    # Create scatter plot
    fig = px.scatter(
        df_filtered,
        x='معدل دوران الكميات',
        y='معدل دوران الفواتير',
        size='الشراء المقترح',
        color='الأولوية',
        hover_data=['اسم المنتج', 'الكمية المتاحة'],
        title="تحليل مقارن لمعدلات الدوران",
        labels={
            'معدل دوران الكميات': 'معدل دوران الكميات',
            'معدل دوران الفواتير': 'معدل دوران الفواتير'
        },
        color_discrete_map={
            'عاجل جداً': '#FF4444',
            'عاجل': '#FF8800', 
            'متوسط': '#FFAA00',
            'منخفض': '#00AA00'
        }
    )
    
    fig.update_layout(
        xaxis_title="معدل دوران الكميات",
        yaxis_title="معدل دوران الفواتير",
        height=500
    )
    
    return fig

def create_priority_chart(df):
    """Create priority analysis chart"""
    priority_counts = df['الأولوية'].value_counts()
    
    colors = {
        'عاجل جداً': '#FF4444',
        'عاجل': '#FF8800', 
        'متوسط': '#FFAA00',
        'منخفض': '#00AA00'
    }
    
    fig = px.bar(
        x=priority_counts.index,
        y=priority_counts.values,
        title="توزيع المنتجات حسب الأولوية",
        color=priority_counts.index,
        color_discrete_map=colors
    )
    
    fig.update_layout(
        xaxis_title="الأولوية",
        yaxis_title="عدد المنتجات",
        showlegend=False,
        height=400
    )
    
    return fig

def create_stock_days_chart(df):
    """Create stock days distribution chart"""
    # Filter products with purchase recommendations
    df_filtered = df[df['الشراء المقترح'] > 0]
    
    if df_filtered.empty:
        return None
    
    fig = px.histogram(
        df_filtered,
        x='أيام التغطية',
        nbins=20,
        title="توزيع أيام التغطية للمنتجات المطلوب شراؤها",
        labels={'أيام التغطية': 'أيام التغطية', 'count': 'عدد المنتجات'}
    )
    
    fig.update_layout(
        xaxis_title="أيام التغطية",
        yaxis_title="عدد المنتجات",
        height=400
    )
    
    return fig

# Main Streamlit application
def main():
    st.title("🛒 نظام اقتراح المشتريات المتقدم")
    st.markdown("---")
    
    # Sidebar for parameters
    with st.sidebar:
        st.header("⚙️ إعدادات النظام")
        
        target_month = st.selectbox(
            "اختر الشهر المستهدف", 
            options=list(range(1, 13)),
            index=datetime.now().month - 1,
            format_func=lambda x: [
                "يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
                "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"
            ][x-1]
        )
        
        target_year = st.number_input(
            "أدخل السنة المستهدفة", 
            value=datetime.now().year, 
            min_value=2020, 
            max_value=2030
        )
        
        safety_stock_days = st.slider(
            "أيام مخزون الأمان",
            min_value=7,
            max_value=90,
            value=30,
            help="عدد الأيام الإضافية للحماية من نفاد المخزون"
        )
        
        st.markdown("---")
        st.markdown("### 📊 معلومات النظام")
        st.info("""
        **يتم حساب الشراء بناءً على:**
        - الشهر السابق للشهر المختار
        - 3 شهور مقابلة في السنة السابقة  
        - متوسط المبيعات الشهرية
        - مخزون الأمان
        - معدل دوران المخزون
        - تحليل الأولوية
        """)
    
    # Load data
    with st.spinner("جاري تحميل البيانات..."):
        sales, stock, purchases = load_data()
    
    if sales is None or stock is None or purchases is None:
        st.error("❌ فشل في تحميل البيانات. يرجى التأكد من وجود الملفات المطلوبة.")
        return
    
    # Display data summary
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("📈 سجلات المبيعات", f"{len(sales):,}")
    with col2:
        st.metric("📦 منتجات المخزون", f"{len(stock):,}")
    with col3:
        st.metric("🛍️ سجلات المشتريات", f"{len(purchases):,}")
    
    st.markdown("---")
    
    # Generate plan button
    if st.button("🔄 توليد خطة الشراء", type="primary", use_container_width=True):
        with st.spinner("جاري توليد خطة الشراء..."):
            try:
                plan = generate_plan(sales, stock, purchases, target_month, target_year, safety_stock_days)
                
                if plan.empty:
                    st.warning("⚠️ لم يتم إنشاء خطة شراء. تحقق من البيانات.")
                    return
                
                st.success("✅ تم توليد خطة الشراء بنجاح!")
                
                # Key metrics
                st.subheader("📊 مؤشرات رئيسية")
                
                total_recommended = plan['الشراء المقترح'].sum()
                products_to_buy = len(plan[plan['الشراء المقترح'] > 0])
                avg_quantity_turnover = plan['معدل دوران الكميات'].mean()
                avg_invoice_turnover = plan['معدل دوران الفواتير'].mean()
                urgent_products = len(plan[plan['الأولوية'].isin(['عاجل', 'عاجل جداً'])])
                
                col1, col2, col3, col4, col5 = st.columns(5)
                with col1:
                    st.metric("إجمالي القطع المقترحة", f"{total_recommended:,.0f}")
                with col2:
                    st.metric("المنتجات المطلوب شراؤها", f"{products_to_buy:,}")
                with col3:
                    st.metric("متوسط دوران الكميات", f"{avg_quantity_turnover:.2f}")
                with col4:
                    st.metric("متوسط دوران الفواتير", f"{avg_invoice_turnover:.2f}")
                with col5:
                    st.metric("المنتجات العاجلة", f"{urgent_products:,}")
                
                # Charts
                st.subheader("📈 التحليل البصري")
                
                # Turnover analysis charts
                turnover_charts = create_turnover_charts(plan)
                st.plotly_chart(turnover_charts, use_container_width=True)
                
                col1, col2 = st.columns(2)
                with col1:
                    priority_chart = create_priority_chart(plan)
                    st.plotly_chart(priority_chart, use_container_width=True)
                
                with col2:
                    # Combined analysis chart
                    combined_chart = create_combined_analysis_chart(plan)
                    if combined_chart:
                        st.plotly_chart(combined_chart, use_container_width=True)
                
                # Stock days distribution
                stock_chart = create_stock_days_chart(plan)
                if stock_chart:
                    st.plotly_chart(stock_chart, use_container_width=True)
                
                # Filters
                st.subheader("🔍 تصفية النتائج")
                
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    quantity_turnover_filter = st.multiselect(
                        "تصنيف دوران الكميات:",
                        options=plan['تصنيف دوران الكميات'].unique(),
                        default=plan['تصنيف دوران الكميات'].unique()
                    )
                
                with col2:
                    invoice_turnover_filter = st.multiselect(
                        "تصنيف دوران الفواتير:",
                        options=plan['تصنيف دوران الفواتير'].unique(),
                        default=plan['تصنيف دوران الفواتير'].unique()
                    )
                
                with col3:
                    priority_filter = st.multiselect(
                        "الأولوية:",
                        options=plan['الأولوية'].unique(),
                        default=plan['الأولوية'].unique()
                    )
                
                with col4:
                    min_purchase = st.number_input(
                        "الحد الأدنى للشراء المقترح:",
                        min_value=0.0,
                        value=0.0,
                        step=1.0
                    )
                
                # Apply filters
                filtered_plan = plan[
                    (plan['تصنيف دوران الكميات'].isin(quantity_turnover_filter)) &
                    (plan['تصنيف دوران الفواتير'].isin(invoice_turnover_filter)) &
                    (plan['الأولوية'].isin(priority_filter)) &
                    (plan['الشراء المقترح'] >= min_purchase)
                ]
                
                st.subheader(f"📋 خطة الشراء ({len(filtered_plan)} منتج)")
                
                # Display plan
                st.dataframe(
                    filtered_plan,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "الأولوية": st.column_config.SelectboxColumn(
                            "الأولوية",
                            options=["عاجل جداً", "عاجل", "متوسط", "منخفض"],
                        ),
                        "الشراء المقترح": st.column_config.NumberColumn(
                            "الشراء المقترح",
                            format="%.0f"
                        ),
                        "معدل الدوران": st.column_config.NumberColumn(
                            "معدل الدوران",
                            format="%.2f"
                        )
                    }
                )
                
                # Summary by category
                st.subheader("📋 ملخص حسب الفئة")
                
                if 'فئة المنتج' in filtered_plan.columns:
                    category_summary = filtered_plan.groupby('فئة المنتج').agg({
                        'الشراء المقترح': 'sum',
                        'الباركود': 'count'
                    }).round(2)
                    category_summary.columns = ['إجمالي الشراء المقترح', 'عدد المنتجات']
                    category_summary = category_summary.sort_values('إجمالي الشراء المقترح', ascending=False)
                    st.dataframe(category_summary, use_container_width=True)
                
                # Download options
                st.subheader("📥 تحميل النتائج")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    excel_file = to_excel(filtered_plan, "purchase_plan")
                    st.download_button(
                        label="📊 تحميل خطة الشراء (Excel)",
                        data=excel_file,
                        file_name=f"purchase_plan_{target_year}_{target_month:02d}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                with col2:
                    # Urgent products only
                    urgent_products = filtered_plan[filtered_plan['الأولوية'].isin(['عاجل', 'عاجل جداً'])]
                    if not urgent_products.empty:
                        urgent_excel = to_excel(urgent_products, "urgent_purchases")
                        st.download_button(
                            label="🚨 تحميل المنتجات العاجلة فقط",
                            data=urgent_excel,
                            file_name=f"urgent_purchases_{target_year}_{target_month:02d}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                
            except Exception as e:
                st.error(f"❌ خطأ في توليد خطة الشراء: {str(e)}")
                st.exception(e)

if __name__ == "__main__":
    main()
